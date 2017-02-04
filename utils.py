"""Classes and methods for Central Line Event Calculator analysis"""

from tkinter import filedialog, messagebox
from openpyxl import Workbook, load_workbook, cell
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, colors

from datetime import datetime, timedelta, date

import functools
import string
import os

# define options for opening or saving a file
OPTIONS = {}
OPTIONS['defaultextension'] = '.xlsx'
OPTIONS['filetypes'] = [('Excel files', '.xlsx .xls'), ('all files', '.*')]

def error_message(title, message):
    messagebox.showwarning(title, message)

def get_file_path(title):
    """Opens TkInter dialogue window and returns user specefied file path."""
    OPTIONS['title'] = title
    path = filedialog.askopenfilename(**OPTIONS)
    return path

def get_file_directory(title):
    """Opens TkInter dialogue window and returns user specefied file directory."""
    OPTIONS['title'] = title
    directory = filedialog.askdirectory()
    return directory

def verify_paths(path1, path2, path3, path4, output):
    """Verifies all file paths are Excel files with the correct data formats."""
    first_valid = verify_excel_file(path1)
    second_valid = verify_excel_file(path2)
    third_valid = verify_excel_file(path3)
    fourth_valid = verify_excel_file(path4)
    if not first_valid or not second_valid or not third_valid or not fourth_valid:
        error_message("Invalid Data Supplied", "All input data must be Excel files.")
        return False
    if not file_exits(path1) or not file_exits(path2) or not file_exits(path3) \
    or not file_exits(path4):
        error_message(
            "Invalid Data Supplied",
            "One or more of the specified input files does not exist.\n" + \
            "Check the file locations and names and try again.")
        return False
    return True
        
def file_exits(path):
    return os.path.isfile(path)

def verify_excel_file(path):
    """Ensures that a path is an Excel file."""
    if path == '':
        return False
    try:
        file_ending = path.split('.')[1]
    except IndexError:
        return False
    if file_ending == 'xlsx' or file_ending == 'xls':
        return True
    return False

def process_data(title, admit_path, line_path, clabsi_path,  clanc_path, out_path, start_range, end_range):
    """Read in each file and writes results to the out_path."""
    events = {}
    print("processing...")
    patients = read_line_data(line_path, start_range, end_range)
    print("processing...")
    read_patient_data(admit_path, patients, start_range, end_range)
    print("processing...")
    read_clabsi_data(clabsi_path, patients)
    print("processing...")
    read_clanc_data(clanc_path, patients)
    print("processing...")
    generate_patient_output(title, out_path, patients, events, start_range, end_range)
    print("processing...")
    generate_line_output(title, out_path, patients, events)
    return True

def read_line_data(path, start_range, end_range):
    """Read in line data. Stores lines as Line objects associated with Patient IDs."""
    work_book = load_workbook(path, read_only=True)
    w_sheet = work_book.active
    patients = {}
    index = 2
    while index <= w_sheet.max_row and w_sheet['A' + str(index)].value is not None:
        p_id = w_sheet['A' + str(index)].value
        line_id = w_sheet['B' + str(index)].value
        line_type = w_sheet['C' + str(index)].value
        lumens = w_sheet['D' + str(index)].value
        in_date = w_sheet['E' + str(index)].value
        out_date = w_sheet['F' + str(index)].value
        if out_date is None:
            out_date = w_sheet['G' + str(index)].value
        removal_reason = w_sheet['H' + str(index)].value

        # Spreadsheet format check
        # if not isinstance(p_id, int):
        #     raise BadFormatException("Patient ID Numbers in Column A of Line Data must be numbers.")
        if not isinstance(line_id, int):
            raise BadFormatException("Line ID Numbers in Column B of Line Data must be numbers.")
        if not isinstance(line_type, str):
            raise BadFormatException("Line Types in Column C of Line Data must be text.")
        if not isinstance(lumens, int):
            raise BadFormatException("Lumen Count in Column D of Line Data must be numbers.")
        if not isinstance(in_date, datetime):
            raise BadFormatException("Patient Admission Dates in Column E of Patient Data must be dates.")
        if not isinstance(in_date, datetime):
            raise BadFormatException("Patient Disscharge Dates in Column F and G of Patient Data must be dates.")
        if not isinstance(removal_reason, str) and removal_reason is not None:
            raise BadFormatException("Reason For Removal in Column H of Line Data must be text.")
        
        #Check Dates
        if (in_date < start_range and out_date < start_range) or (in_date > end_range):
            index += 1
            continue #Do not add dates outside of range 
        #commented out because i think it confuses the end result.
        # elif in_date < start_range and out_date > start_range:
        #     in_date = start_range
        # elif in_date < end_range and out_date > end_range:
        #     out_date = end_range

        if not p_id in patients:
            p = Patient(p_id)
            patients[p_id] = p

        l = Line(line_id, line_type, lumens, in_date, out_date, removal_reason, start_range, end_range)   
        patients[p_id].add_line(l)
        index += 1
    return patients


def read_patient_data(path, patients, start_range, end_range):
    """Read in patient admit data. Returns a dictionary of Patient objects (Key: ID Number)."""
    work_book = load_workbook(path, read_only=True)
    w_sheet = work_book.active
    index = 2
    while index <= w_sheet.max_row:
        p_id = w_sheet['A' + str(index)].value
        in_date = w_sheet['B' + str(index)].value
        out_date = w_sheet['C' + str(index)].value

        # Spreadsheet format check
        if not isinstance(p_id, int):
            raise BadFormatException("Patient ID Numbers in Column A of Patient Data must be numbers.")
        if not isinstance(in_date, datetime):
            raise BadFormatException("Patient Admission Dates in Column B of Patient Data must be dates.")
        if not isinstance(in_date, datetime):
            raise BadFormatException("Patient Disscharge Dates in Column C of Patient Data must be dates.")

        if out_date < start_range or in_date > end_range:
            index += 1
            continue

        if p_id in patients and check_full_day_admit(in_date, out_date):
            patients[p_id].add_visit(Visit(patients[p_id], in_date, out_date))

        index += 1

def read_clabsi_data(path, patients):
    work_book = load_workbook(path, read_only=True)
    w_sheet = work_book.active
    index = 2
    while index <= w_sheet.max_row:
        p_id = w_sheet['A' + str(index)].value
        clabsi_date = w_sheet['B' + str(index)].value

        if not isinstance(p_id, int):
            raise BadFormatException("Patient ID Numbers in Column A of CLABSI Data must be numbers.")
        if not isinstance(clabsi_date, datetime):
            raise BadFormatException("CLABSI Date in Column B of CLABSI Data must be a date.")

        if p_id in patients:    
            p = patients[p_id]
            lines = []
            for l in p.lines:
                if l.in_date <= clabsi_date and l.out_date >= clabsi_date:
                    lines.append(l)

            event = CLABSI(p, lines, clabsi_date)
            for visit in p.visits:
                if clabsi_date > visit.check_in_date and clabsi_date <= visit.check_out_date + timedelta(days=1):
                    event.inpatient = True
                else:
                    event.inpatient = False

            p.clabsis.append(event)

       
        index += 1


def read_clanc_data(path, patients):
    work_book = load_workbook(path, read_only=True)
    w_sheet = work_book.active
    index = 2
    while index <= w_sheet.max_row:
        p_id = w_sheet['A' + str(index)].value
        line_id = w_sheet['B' + str(index)].value
        clanc_date = w_sheet['C' + str(index)].value

        if not isinstance(p_id, int):
            raise BadFormatException("Patient ID Numbers in Column A of CLANC Data must be numbers.")
        if not isinstance(line_id, int):
            raise BadFormatException("Line ID Numbers in Column B of CLANC Data must be numbers.")
        if not isinstance(clanc_date, datetime):
            raise BadFormatException("CLABSI Date in Column C of CLANC Data must be a date.")
        
        if p_id in patients:
            p = patients[p_id]
            line = [l for l in p.lines if line_id == l.line_id]
            if line:
                line = line[0]
            else:
                index += 1
                continue
            event = CLANC(p, line, clanc_date)
            for visit in p.visits:
                if clanc_date >= visit.check_in_date and clanc_date <= visit.check_out_date:
                    event.inpatient = True
                else:
                    event.inpatient = False

            p.clancs.append(event)
            line.clanc = event
        index += 1


def check_full_day_admit(in_time, out_time):
    """Returns True if Patient was admitted for at least 24 hours."""
    diff = out_time - in_time
    if diff.days == 0:
        return False
    return True

def generate_patient_output(title, path, patients, events, start_range, end_range):
    """Writes patient-only analysis to new Excel file."""
    work_book = Workbook()
    w_sheet = work_book.active

    #Column Titles
    w_sheet.title = 'Output Individual Patient'
    w_sheet['A1'] = 'Patient ID'
    w_sheet['B1'] = 'Total Lines'
    w_sheet['C1'] = 'Sum of all Line Days'
    w_sheet['D1'] = 'Inpatient Line Days'
    w_sheet['E1'] = 'Outpatient Line Days'
    w_sheet['F1'] = 'Mean Duration of Line (Days)'
    w_sheet['G1'] = 'Total Days with any Catheter'
    w_sheet['H1'] = 'Catheter Density (Sum of all Line Days/Total Days with any catheter)'
    w_sheet['I1'] = 'Sum of all Lumen Days'
    w_sheet['J1'] = 'Inpatient Lumen Days'
    w_sheet['K1'] = 'Outpatient Lumen Days'
    w_sheet['L1'] = 'Inpatient Lumen Density (Inpatient Lumen Days/Total Inpatient Days With A Catheter)'
    w_sheet['M1'] = 'Outpatient Lumen Density (Outpatient Lumen Days/Total Outpatient Days With A Catheter)'
    w_sheet['N1'] = 'Total Lumen Density (Sum of all Lumen Days/Total Days with any cather)'
    w_sheet['O1'] = 'CLABSIs'
    w_sheet['P1'] = 'Inpatient CLABSIs'
    w_sheet['Q1'] = 'Outpatient CLABSIs'
    w_sheet['R1'] = 'Inpatient CLABSI Rate (x1000)'
    w_sheet['S1'] = 'Outpatient CLABSI Rate (x1000)'
    w_sheet['T1'] = "CLABSI Rate (x1000)"
    w_sheet['U1'] = 'CLANCs'
    w_sheet['V1'] = 'Inpatient CLANCs'
    w_sheet['W1'] = 'Outpatient CLANCs'
    w_sheet['X1'] = "Inpatient CLANC Rate (x1000)"
    w_sheet['Y1'] = "Outpatient CLANC Rate (x1000)"
    w_sheet['Z1'] = "CLANC Rate (x1000)"
    w_sheet['AA1'] = "ALL EVENT Rate (x1000)"
    w_sheet['AB1'] = "Inpatient Catheter Days"
    w_sheet['AC1'] = "Outpatient Catheter Days"

    pop_inp = 0
    pop_out = 0

    row = 2
    for p_id in patients:
        p = patients[p_id]
        calculate_inpatient_line_days(p)
        for l in p.lines:
            p.inpatient_lumen_time += l.inpatient_lumen_time

        in_clabsi = 0
        out_clabsi = 0
        in_clanc = 0
        out_clanc = 0
        for e in p.clabsis:
            if e.inpatient:
                in_clabsi += 1
            else:
                out_clabsi += 1
        for e in p.clancs:
            if e.inpatient:
                in_clanc += 1
            else:
                out_clanc += 1

        total_cath_days, inp_cath_days, outp_cath_days = calculate_total_cath_days(p, start_range, end_range) if p.lines else 0
        pop_inp += inp_cath_days
        pop_out += outp_cath_days
        print(total_cath_days, inp_cath_days, outp_cath_days)

        w_sheet['A' + str(row)] = p_id
        w_sheet['B' + str(row)] = len(p.lines)
        w_sheet['C' + str(row)] = p.total_line_time.days
        w_sheet['D' + str(row)] = p.inpatient_line_time.days
        w_sheet['E' + str(row)] = p.total_line_time.days - p.inpatient_line_time.days
        w_sheet['F' + str(row)] = (p.total_line_time.days/len(p.lines)) if p.lines else 0
        w_sheet['G' + str(row)] = total_cath_days
        w_sheet['H' + str(row)] = (w_sheet['C' + str(row)].value / w_sheet['G' + str(row)].value) if w_sheet['G' + str(row)].value != 0 else 0
        w_sheet['I' + str(row)] = p.total_lumen_time.days #check
        w_sheet['J' + str(row)] = p.inpatient_lumen_time.days
        w_sheet['K' + str(row)] = p.total_lumen_time.days - p.inpatient_lumen_time.days
        w_sheet['L' + str(row)] = (w_sheet['J' + str(row)].value / inp_cath_days) if inp_cath_days else 0 
        w_sheet['M' + str(row)] = (w_sheet['K' + str(row)].value / outp_cath_days) if outp_cath_days else 0 
        w_sheet['N' + str(row)] = (w_sheet['I' + str(row)].value / w_sheet['G' + str(row)].value) if w_sheet['G' + str(row)].value != 0 else 0 
        w_sheet['O' + str(row)] = len(p.clabsis)
        w_sheet['P' + str(row)] = in_clabsi
        w_sheet['Q' + str(row)] = out_clabsi
        w_sheet['R' + str(row)] = ((in_clabsi / inp_cath_days) * 1000) if inp_cath_days else 0
        w_sheet['S' + str(row)] = ((out_clabsi / outp_cath_days) * 1000) if outp_cath_days else 0
        w_sheet['T' + str(row)] = (w_sheet['O' + str(row)].value / w_sheet['G' + str(row)].value * 1000) if w_sheet['G' + str(row)].value != 0 else 0
        w_sheet['U' + str(row)] = len(p.clancs)
        w_sheet['V' + str(row)] = in_clanc
        w_sheet['W' + str(row)] = out_clanc
        w_sheet['X' + str(row)] = ((in_clanc / inp_cath_days) * 1000) if inp_cath_days else 0
        w_sheet['Y' + str(row)] = ((out_clanc / outp_cath_days) * 1000) if outp_cath_days else 0
        w_sheet['Z' + str(row)] = (w_sheet['U' + str(row)].value / w_sheet['G' + str(row)].value * 1000) if w_sheet['G' + str(row)].value != 0 else 0
        w_sheet['AA' + str(row)] = ((in_clanc + out_clanc + in_clabsi + out_clabsi) / w_sheet['G' + str(row)].value * 1000) if w_sheet['G' + str(row)].value != 0 else 0
        w_sheet['AB' + str(row)] = inp_cath_days
        w_sheet['AC' + str(row)] = outp_cath_days
        
        row += 1    

    #Summation Data
    max_index = str(row)
    bottom = row - 1
    w_sheet['A' + max_index] = 'Population Total'
    w_sheet['B' + max_index] = '=SUM(B2:B' + str(bottom) + ')'
    w_sheet['C' + max_index] = '=SUM(C2:C' + str(bottom) + ')'
    w_sheet['D' + max_index] = '=SUM(D2:D' + str(bottom) + ')'
    w_sheet['E' + max_index] = '=SUM(E2:E' + str(bottom) + ')'
    w_sheet['F' + max_index] = '=C' + max_index + '/B' + max_index
    w_sheet['G' + max_index] = '=SUM(G2:G' + str(bottom) + ')'
    w_sheet['H' + max_index] = '=C'+ max_index + '/G' + max_index
    w_sheet['I' + max_index] = '=SUM(I2:I' +  str(bottom) + ')'
    w_sheet['J' + max_index] = '=SUM(J2:J' +  str(bottom) + ')'
    w_sheet['K' + max_index] = '=SUM(K2:K' +  str(bottom) + ')'
    w_sheet['L' + max_index] = '=J'+ max_index + '/' + str(pop_inp)
    w_sheet['M' + max_index] = '=K'+ max_index + '/' + str(pop_out)
    w_sheet['N' + max_index] = '=I'+ max_index + '/G' + max_index
    w_sheet['O' + max_index] = '=SUM(O2:O' +  str(bottom) + ')'
    w_sheet['P' + max_index] = '=SUM(P2:P' +  str(bottom) + ')'
    w_sheet['Q' + max_index] = '=SUM(Q2:Q' +  str(bottom) + ')'
    w_sheet['R' + max_index] = '=P'+ max_index + '/' + str(pop_inp) + "* 1000"
    w_sheet['S' + max_index] = '=Q'+ max_index + '/' + str(pop_out) + "* 1000"
    w_sheet['T' + max_index] = '=O'+ max_index + '/G' + max_index + "* 1000"
    w_sheet['U' + max_index] = '=SUM(U2:U' +  str(bottom) + ')'
    w_sheet['V' + max_index] = '=SUM(V2:V' +  str(bottom) + ')'
    w_sheet['W' + max_index] = '=SUM(W2:W' +  str(bottom) + ')'
    w_sheet['X' + max_index] = '=V'+ max_index + '/' + str(pop_inp) + "* 1000"
    w_sheet['Y' + max_index] = '=W'+ max_index + '/' + str(pop_out) + "* 1000"
    w_sheet['Z' + max_index] = '=U'+ max_index + '/G' + max_index + "* 1000"
    w_sheet['AA' + max_index] = '=(O'+ max_index + '+ U' + max_index + ')/G' + max_index + "* 1000"
    w_sheet['AB' + max_index] = '=SUM(AB2:AB' +  str(bottom) + ')'
    w_sheet['AC' + max_index] = '=SUM(AC2:AC' +  str(bottom) + ')'

    #adjust cell width for titles
    index = 1
    for col in w_sheet.columns:
        w_sheet.column_dimensions[get_column_letter(index)].width = len(col[0].value)
        col[-1].fill =  PatternFill(start_color='008000', end_color='008000', fill_type='solid')
        index += 1

    c = w_sheet['A2']
    w_sheet.freeze_panes = c
    work_book.save(path + "/" + title + " - Output Individual Patient.xlsx")

def generate_line_output(title, path, patients, events):
    """Writes line-only analysis to new Excel file."""
    work_book = Workbook()
    w_sheet = work_book.active

     #Column Titles
    w_sheet.title = 'Output Individual Line'
    w_sheet['A1'] = 'Line ID'
    w_sheet['B1'] = 'Patient ID'
    w_sheet['C1'] = 'Number of Lumens'
    w_sheet['D1'] = 'Date of Insertion (or first evaluation)'
    w_sheet['E1'] = 'Date of Removal (or last evalulation)'
    w_sheet['F1'] = 'Line Days (any catheter)'
    w_sheet['G1'] = 'Inpatient Line Days (any catheter)'
    w_sheet['H1'] = 'Outpatient Line Days (any catheter)'
    w_sheet['I1'] = 'Lumen Days (Line Days x Number of Lumens)'
    w_sheet['J1'] = 'Inpatient Lumen Days (Line Days x Number of Lumens)'
    w_sheet['K1'] = 'Outpatient Lumen Days (Line Days x Number of Lumens)'

    w_sheet['L1'] = "Number of Inpatient CLABSIs"
    w_sheet['M1'] = "Number of Outpatient CLABSIs"
    w_sheet['N1'] = "Total CLABSIs"
    w_sheet['O1'] =  "Number of Inpatient CLANCs"
    w_sheet['P1'] =  "Number of Outpatient CLANCs"
    w_sheet['Q1'] =  "Total CLANCs"
    w_sheet['R1'] =  "Time from CLANC to line removal (Days)"
    w_sheet['S1'] =  "Reason For Line Removal"

    w_sheet['T1'] =  "ALL EVENTS"
    w_sheet['U1'] =  "ALL EVENT RATE (x1000)"

    w_sheet['V1'] =  "Inpatient CLASBI Rate (x1000)"
    w_sheet['W1'] =  "Outpatient CLASBI Rate (x1000)"
    w_sheet['X1'] =  "Total CLASBI Rate (x1000)"
    w_sheet['Y1'] =  "Inpatient CLANC Rate (x1000)"
    w_sheet['Z1'] =  "Outpatient CLANC Rate (x1000)"
    w_sheet['AA1'] =  "Total CLANC Rate (x1000)"

    row = 2
    for p_id in patients:
        p = patients[p_id]
        for l in p.lines:
            w_sheet['A' + str(row)] = l.line_id
            w_sheet['B' + str(row)] = p_id
            w_sheet['C' + str(row)] = l.lumens
            w_sheet['D' + str(row)] = l.in_date
            w_sheet['E' + str(row)] = l.out_date
            w_sheet['F' + str(row)] = l.total_time.days
            w_sheet['G' + str(row)] = l.inpatient_line_time.days
            w_sheet['H' + str(row)] = l.total_time.days - l.inpatient_line_time.days
            w_sheet['I' + str(row)] = l.lumen_days.days
            w_sheet['J' + str(row)] = l.inpatient_lumen_time.days
            w_sheet['K' + str(row)] = (l.lumen_days - l.inpatient_lumen_time).days

            num_inpatient = 0
            num_outpatient = 0
            for e in p.clabsis:
                if l in e.lines:
                    if e.inpatient:
                        num_inpatient += 1/(len(e.lines))
                    else:
                        num_outpatient += 1/(len(e.lines))

            w_sheet['L' + str(row)] = num_inpatient
            w_sheet['M' + str(row)] = num_outpatient
            w_sheet['N' + str(row)] = num_inpatient + num_outpatient

            num_in_clancs = 0
            num_out_clancs = 0
            if l.clanc:
                if l.clanc.inpatient:
                    num_in_clancs = 1
                else:
                    num_out_clancs = 1
                diff = l.out_date - l.clanc.date
                w_sheet['R' + str(row)] = diff.days
            else:
                 w_sheet['R' + str(row)] = "No CLANC Reported"
            w_sheet['O' + str(row)] = num_in_clancs
            w_sheet['P' + str(row)] = num_out_clancs
            w_sheet['Q' + str(row)] = num_in_clancs + num_out_clancs

            w_sheet['S' + str(row)] = l.removal_reason

            total_events = num_in_clancs + num_out_clancs + num_inpatient + num_outpatient
            w_sheet['T' + str(row)] = total_events
            w_sheet['U' + str(row)] = ((total_events / l.total_time.days) * 1000) if l.total_time.days else 0
            

            w_sheet['X' + str(row)] = (((num_inpatient + num_outpatient) / l.total_time.days) * 1000) if l.total_time.days else 0
            w_sheet['AA' + str(row)] = (((num_in_clancs + num_out_clancs) / l.total_time.days) * 1000) if l.total_time.days else 0
            
            #clasbi in/out rate
            w_sheet['V' + str(row)] = ((num_inpatient / l.inpatient_line_time.days) * 1000) if l.inpatient_line_time.days else 0
            w_sheet['W' + str(row)] = ((num_outpatient / (l.total_time.days - l.inpatient_line_time.days)) * 1000) if (l.total_time.days - l.inpatient_line_time.days) else 0

            #clanc in/out rate
            w_sheet['Y' + str(row)] = ((num_in_clancs / l.inpatient_line_time.days) * 1000) if l.inpatient_line_time.days else 0
            w_sheet['Z' + str(row)] = ((num_out_clancs / (l.total_time.days - l.inpatient_line_time.days)) * 1000) if (l.total_time.days - l.inpatient_line_time.days) else 0
            w_sheet['D' + str(row)].number_format = 'dd-mmm-yy'
            w_sheet['E' + str(row)].number_format = 'dd-mmm-yy'

            row += 1

    #adjust cell width for titles
    index = 1
    for col in w_sheet.columns:
        w_sheet.column_dimensions[get_column_letter(index)].width = max(10, len(col[0].value))
        index += 1

    c = w_sheet['A2']
    w_sheet.freeze_panes = c
    work_book.save(path + "/" + title + " - Output Individual Line.xlsx")

def calculate_total_cath_days(p, start_range, end_range):
    """Returns the total number of days a Patient has ANY catheter."""
    sorted(p.lines)
    lines_in_range = []
    for l in p.lines:
        if l.out_date < start_range:
            continue
        elif l.in_date > end_range:
            break
        lines_in_range += [l]
    date_range = []
    index = 0
    inpatient_cath_days = []

    for l in lines_in_range:
        start = l.in_date if l.in_date >= start_range else start_range
        end = l.out_date if l.out_date <= end_range else end_range
        date_range += [timedelta(days = d) + start.date() for d in range((end-start).days)]
        for v in p.visits:
            if end < v.check_in_date or start > v.check_out_date:
                continue
            elif v.check_in_date > start and v.check_out_date < end:
                tmp = [timedelta(days = d) + v.check_in_date for d in range((v.check_out_date - v.check_in_date).days)]
                inpatient_cath_days += [date(d.year, d.month, d.day) for d in tmp]
            elif v.check_in_date <= start and v.check_out_date >= end:
                tmp = [timedelta(days = d) + start for d in range((end - start).days)]
                inpatient_cath_days += [date(d.year, d.month, d.day) for d in tmp]
            elif v.check_in_date <= start and v.check_out_date < end:
                tmp = [timedelta(days = d) + start for d in range((v.check_out_date - start).days)]
                inpatient_cath_days += [date(d.year, d.month, d.day) for d in tmp]
            elif v.check_in_date > start and v.check_out_date <= end:
                tmp = [timedelta(days = d) + v.check_in_date for d in range((end - v.check_in_date).days)]
                inpatient_cath_days += [date(d.year, d.month, d.day) for d in tmp]
    inp_cath_days = len(set(inpatient_cath_days))
    total_cath_days = len(set(date_range))
    return [total_cath_days, inp_cath_days, total_cath_days - inp_cath_days]

def calculate_inpatient_line_days(p):
    for v in p.visits:
        for l in p.lines:
            if l.out_date < v.check_in_date or l.in_date > v.check_out_date:
                continue
            elif v.check_in_date > l.in_date and v.check_out_date < l.out_date:
                p.inpatient_line_time += v.check_out_date - v.check_in_date
                l.inpatient_line_time += v.check_out_date - v.check_in_date
            elif v.check_in_date <= l.in_date and v.check_out_date >= l.out_date:
                p.inpatient_line_time += l.out_date - l.in_date
                l.inpatient_line_time += l.out_date - l.in_date
            elif v.check_in_date <= l.in_date and v.check_out_date < l.out_date:
                p.inpatient_line_time += v.check_out_date - l.in_date
                l.inpatient_line_time += v.check_out_date - l.in_date
            elif v.check_in_date > l.in_date and v.check_out_date <= l.out_date:
                p.inpatient_line_time += l.out_date - v.check_in_date
                l.inpatient_line_time += l.out_date - v.check_in_date
            l.inpatient_lumen_time = l.lumens * l.inpatient_line_time

class Patient():
    """Patient Class contains lists of Visits, Lines and a Dictionary of Events."""
    def __init__(self, patient_id):
        self.visits = []
        self.lines = []
        #self.events = {} # possibly becoming depricated. Review for later. Comment meant to trigger pylint.

        #list of clabsis and possible clanc events
        self.clabsis = []
        self.clancs = []

        self.patient_id = patient_id
        self.total_visit_time = timedelta(0)
        self.total_line_time = timedelta(0)
        self.total_lumen_time = timedelta(0)
        self.inpatient_line_time = timedelta(0)
        self.inpatient_lumen_time = timedelta(0)

    def add_visit(self, v):
        """Adds a Visit object to the list of Visits and adds the time to total_visit_time."""
        assert isinstance(v, Visit), "new visits must be of type Visit"
        self.visits.append(v)
        if self.total_visit_time is None:
            self.total_visit_time = v.total_time
        else:
            self.total_visit_time += v.total_time

    def add_line(self, l):
        """Adds a Line object to the list of Visits and adds the time to total_line_time."""
        assert isinstance(l, Line), "new lines must be of type Line"
        self.lines.append(l)
        if self.total_line_time is None:
            self.total_line_time = l.total_time
            self.total_lumen_time = l.lumen_days
        else:
            self.total_line_time += l.total_time
            self.total_lumen_time += l.lumen_days

class Visit():
    """Visit Class stores datetime info for a single Patient Visit."""
    def __init__(self, patient, in_date, out_date):
        self.patient = patient
        self.check_in_date = in_date
        self.check_out_date = out_date
        self.total_time = in_date - out_date

@functools.total_ordering
class Line():
    """Line Class stores data for a single Line in a Patient. Records a Dictionary of Events."""
    def __init__(self, line_id, line_type, lumens, in_date, out_date, removal_reason, start_range, end_range):
        self.line_type = line_type
        self.in_date = in_date
        self.out_date = out_date
        self.line_id = line_id
        self.lumens = lumens
        if start_range > in_date:
            in_date = start_range
        if end_range < out_date:
            out_date = end_range
        self.total_time = out_date - in_date
        self.lumen_days = self.total_time * self.lumens
        self.removal_reason = removal_reason
        self.inpatient_line_time = timedelta(0)
        self.inpatient_lumen_time = timedelta(0)


        #list of clabsis and variable for possible clanc event
        self.clabsis = []
        self.clanc = None

    def __lt__(self, other):
        if self.in_date == other.in_date:
            return self.out_date < other.out_date
        return self.in_date < self.out_date

    def __eq__(self, other):
        return self.in_date == other.in_date and self.out_date == other.out_date

class CLABSI():
    """Class for CLABSI event. used becuase required infectious information is more complicated"""
    def __init__(self, patient, lines, date):
        self.patient = patient
        self.lines = lines
        self.date = date
        self.inpatient = False


class CLANC(): 
    """Class for CLANC event. used becuase required non-infect information is more complicated"""
    def __init__(self, patient, line, date):
        self.patient = patient
        self.line = line
        self.date = date
        self.inpatient = False

class BadFormatException(Exception):
    def __init__(self, value):
        self.parameter = value
    def __str__(self):
        return self.parameter
